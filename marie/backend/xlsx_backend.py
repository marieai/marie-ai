import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from marie.backend.base_backend import DocumentBackend

_log = logging.getLogger(__name__)

# Default cell dimensions in points
_DEFAULT_COL_WIDTH = 64.0  # ~8.43 characters * ~7.6 px
_DEFAULT_ROW_HEIGHT = 15.0


def _get_col_widths(sheet, max_col: int) -> list[float]:
    """Get column widths in points for columns 1..max_col."""
    widths = []
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        dim = sheet.column_dimensions.get(letter)
        if dim and dim.width:
            # openpyxl width is in characters; approximate conversion to points
            widths.append(dim.width * 7.0)
        else:
            widths.append(_DEFAULT_COL_WIDTH)
    return widths


def _get_row_heights(sheet, max_row: int) -> list[float]:
    """Get row heights in points for rows 1..max_row."""
    heights = []
    for r in range(1, max_row + 1):
        dim = sheet.row_dimensions.get(r)
        if dim and dim.height:
            heights.append(dim.height)
        else:
            heights.append(_DEFAULT_ROW_HEIGHT)
    return heights


def _text_to_words(text: str, x0: float, y0: float, x1: float, y1: float) -> list[dict]:
    parts = text.split()
    if not parts:
        return []
    total_chars = sum(len(w) for w in parts)
    if total_chars == 0:
        return []
    span = x1 - x0
    words = []
    cx = x0
    for w in parts:
        frac = len(w) / total_chars
        wx1 = cx + span * frac
        words.append(
            {
                "text": w,
                "bbox": [round(cx, 2), round(y0, 2), round(wx1, 2), round(y1, 2)],
                "confidence": 1.0,
            }
        )
        cx = wx1
    return words


class XlsxBackend(DocumentBackend):
    @classmethod
    def supported_formats(cls) -> set[str]:
        return {"xlsx"}

    @classmethod
    def is_available(cls) -> bool:
        try:
            import openpyxl  # noqa: F401

            return True
        except ImportError:
            return False

    def convert(self, file_path: str, **kwargs) -> dict:
        wb = load_workbook(filename=file_path, data_only=True)

        results: list[dict] = []
        for sheet_idx, name in enumerate(wb.sheetnames):
            sheet = wb[name]

            # Skip chartsheets
            if not hasattr(sheet, "iter_rows"):
                continue

            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            if max_row == 0 or max_col == 0:
                results.append(
                    {
                        "words": [],
                        "lines": [],
                        "meta": {"page": sheet_idx, "width": 0, "height": 0},
                    }
                )
                continue

            col_widths = _get_col_widths(sheet, max_col)
            row_heights = _get_row_heights(sheet, max_row)

            # Precompute cumulative x/y offsets
            x_offsets = [0.0]
            for w in col_widths:
                x_offsets.append(x_offsets[-1] + w)

            y_offsets = [0.0]
            for h in row_heights:
                y_offsets.append(y_offsets[-1] + h)

            page_width = x_offsets[-1]
            page_height = y_offsets[-1]

            # Build set of cells hidden by merges (not the top-left anchor)
            hidden: set[tuple[int, int]] = set()
            merge_spans: dict[tuple[int, int], tuple[int, int]] = {}
            for mr in sheet.merged_cells.ranges:
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        if r == mr.min_row and c == mr.min_col:
                            merge_spans[(r, c)] = (
                                mr.max_row - mr.min_row + 1,
                                mr.max_col - mr.min_col + 1,
                            )
                        else:
                            hidden.add((r, c))

            lines: list[dict] = []
            for row in sheet.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col
            ):
                for cell in row:
                    r, c = cell.row, cell.column
                    if (r, c) in hidden:
                        continue
                    val = cell.value
                    if val is None:
                        continue
                    text = str(val).strip()
                    if not text:
                        continue

                    row_span, col_span = merge_spans.get((r, c), (1, 1))
                    x0 = x_offsets[c - 1]
                    y0 = y_offsets[r - 1]
                    x1 = x_offsets[min(c - 1 + col_span, len(x_offsets) - 1)]
                    y1 = y_offsets[min(r - 1 + row_span, len(y_offsets) - 1)]

                    words = _text_to_words(text, x0, y0, x1, y1)
                    if words:
                        lines.append(
                            {
                                "text": text,
                                "bbox": [
                                    round(x0, 2),
                                    round(y0, 2),
                                    round(x1, 2),
                                    round(y1, 2),
                                ],
                                "words": words,
                            }
                        )

            results.append(
                {
                    "words": [w for line in lines for w in line["words"]],
                    "lines": lines,
                    "meta": {
                        "page": sheet_idx,
                        "width": page_width,
                        "height": page_height,
                    },
                }
            )

        wb.close()
        return {"mode": "parsed", "results": results, "pages": len(results)}
